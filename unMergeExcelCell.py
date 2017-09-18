#!/usr/bin/env python
#-*- coding -*-

import os
import sys
import time
try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries
except ImportError:
    print('You must install "openpyxl" to use this script')
    print('if running Anaconda Python use:  conda install openpyxl')
    print('if running Python use:  pip install openpyxl')    
    exit(-1)
try:
    #used for progress bar
    import tqdm
except ImportError:
    print('You must install "tqdm" to use this script')
    print('if running Anaconda Python use:  conda install tqdm')
    print('if running standard Python use:  pip install tqdm')
    exit(-1)


__doc__ = '''Un-merge excel cell and auto fill with the first cell value in the merged cells

Usage: unMergeExcelCell.py <excel file>

excel file: the path for un-merging excel file
'''


def usage():
    print(__doc__)
    exit(-1)
   
def unMergeExcelCell(path):    
    print('Opening workbook {0}'.format(path))
    if not os.path.exists(path):
        print(("Could not find the excel file: " % path))
        return

    #read workbook into memory
    wb = load_workbook(filename = path, data_only=True)
    print('Workbook loaded')
    
    #setup a couple progress counter(s)
    i = 0
    num_worksheets = len(wb.worksheets)
    
    # loop over each worksheet in the workbook
    for ws in wb.worksheets: 
        #maintain the worksheet counter
        i += 1
        
        print('Worksheet progress "{0}" (Sheet {1} of {2})'.format(ws.title,i,num_worksheets))
        #loop over each range in this worksheet
        for rng in tqdm.tqdm(ws.merged_cell_ranges[:]):

            #range_boundaries->Convert a range string into a tuple of boundaries: (min_col, min_row, max_col, max_row)
            #We'll zip the numbers to their logical names so we end up with a dictionary w/ named values for easier use
            rangeBounds = dict(zip(['MIN_COL', 'MIN_ROW', 'MAX_COL', 'MAX_ROW'], range_boundaries(rng)))

            #Capture the value we want to apply to all unmerged cells
            captured_cell_value = ws.cell(column=rangeBounds.get('MIN_COL'),row=rangeBounds.get('MIN_ROW')).value

            #Unmerge the range of cells
            ws.unmerge_cells(rng)
            
            #Update the individual cells to the orignal value
            for row in range(rangeBounds.get('MIN_ROW'),rangeBounds.get('MAX_ROW')+1):
                for col in range(rangeBounds.get('MIN_COL'),rangeBounds.get('MAX_COL')+1):
                    #print("--R=",row," C=",col )
                    ws.cell(column=col, row=row, value=captured_cell_value)
            
    # save the un-merged excel file
    (origin_file, ext) = os.path.splitext(path)
    unmerge_excel_file = origin_file + '_unmerged' + ext
    print('Saving to disk as new file named "{0}"'.format(unmerge_excel_file))
    wb.save(unmerge_excel_file)

    print("Save complete")
    
def main():
    if 2 != len(sys.argv):
        usage()

    excel_file = sys.argv[1]
    unMergeExcelCell(excel_file)

if __name__ == '__main__':
    main()
